from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import re
import statistics

import pandas as pd
from openpyxl import load_workbook


EXPOSURE_BLOCKS = [
    (1, 1),   # A1
    (7, 1),   # A7
    (13, 1),  # A13
    (19, 1),  # A19
    (1, 9),   # I1
    (7, 9),   # I7
    (13, 9),  # I13
    (19, 9),  # I19
]


def exposure_to_seconds(label: str) -> int:
    """Convert Excel labels such as 0\", 15\", 30\", 1', 2' to seconds."""
    text = str(label).strip()
    if text.endswith('"'):
        return int(float(text[:-1]))
    if text.endswith("'"):
        return int(float(text[:-1]) * 60)
    lower = text.lower().replace(" ", "")
    if lower in {"control", "ctrl", "0", "0s", "0sec", "0seconds"}:
        return 0
    m = re.match(r"^(\d+(?:\.\d+)?)(s|sec|secs|second|seconds)$", lower)
    if m:
        return int(float(m.group(1)))
    m = re.match(r"^(\d+(?:\.\d+)?)(m|min|mins|minute|minutes)$", lower)
    if m:
        return int(float(m.group(1)) * 60)
    raise ValueError(f"Cannot parse exposure label: {label!r}")


def exposure_pretty(seconds: int) -> str:
    if seconds == 0:
        return "Control"
    if seconds < 60:
        return f"Treat:{seconds}s"
    if seconds % 60 == 0:
        return f"Treat:{seconds // 60}min"
    return f"Treat:{seconds / 60:g}min"


def time_to_hours(label: str) -> int:
    text = str(label).strip().lower().replace(" ", "")
    text = text.replace("hours", "h").replace("hour", "h")
    if text.endswith("h"):
        return int(float(text[:-1]))
    return int(float(text))


@dataclass(frozen=True)
class TargetBlock:
    cell_line: str
    exposure_label_raw: str
    exposure_seconds: int
    time_h: int
    n1: float | None
    n2: float | None
    n3: float | None
    n4: float | None
    mean_sheet: float | None
    sd_sheet: float | None
    mean_recomputed: float | None
    sd_recomputed: float | None
    mean_t0_normalized: float | None
    sd_t0_normalized: float | None


def _numeric_or_none(x):
    if x is None:
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def read_cap_excel_long(path: str | Path, *, recompute_mean: bool = True) -> pd.DataFrame:
    """
    Read the Gorjet 0--72 h CCA/PDAC workbook and return a long-format table.

    The workbook is organized as 8 small blocks per cell-line sheet:
    0\", 15\", 30\", 1', 2', 3', 4', 5'.  Double quotes are seconds;
    single quotes are minutes. This function recomputes means and sample SDs
    from N=1..N=4 by default because some sheet MEAN cells can be inconsistent
    with the replicate values.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    records: list[dict] = []

    for sheet_name in wb.sheetnames:
        if "(FINAL)" not in sheet_name:
            continue
        ws = wb[sheet_name]
        cell_line = sheet_name.replace("(FINAL)", "").strip()

        for header_row, header_col in EXPOSURE_BLOCKS:
            raw_label = ws.cell(header_row, header_col).value
            if raw_label is None:
                continue
            try:
                exposure_seconds = exposure_to_seconds(str(raw_label))
            except ValueError:
                continue

            block_rows = []
            for offset in range(1, 5):
                row = header_row + offset
                time_raw = ws.cell(row, header_col).value
                if time_raw is None:
                    continue
                try:
                    time_h = time_to_hours(str(time_raw))
                except ValueError:
                    continue

                reps = [_numeric_or_none(ws.cell(row, header_col + j).value) for j in range(1, 5)]
                reps_valid = [x for x in reps if x is not None]
                mean_sheet = _numeric_or_none(ws.cell(row, header_col + 5).value)
                sd_sheet = _numeric_or_none(ws.cell(row, header_col + 6).value)
                mean_recomputed = sum(reps_valid) / len(reps_valid) if reps_valid else None
                sd_recomputed = statistics.stdev(reps_valid) if len(reps_valid) >= 2 else None
                block_rows.append({
                    "cell_line": cell_line,
                    "exposure_label_raw": str(raw_label),
                    "exposure_seconds": exposure_seconds,
                    "exposure_minutes": exposure_seconds / 60.0,
                    "exposure_label": exposure_pretty(exposure_seconds),
                    "time_h": time_h,
                    "n1": reps[0],
                    "n2": reps[1],
                    "n3": reps[2],
                    "n4": reps[3],
                    "mean_sheet": mean_sheet,
                    "sd_sheet": sd_sheet,
                    "mean_recomputed": mean_recomputed,
                    "sd_recomputed": sd_recomputed,
                })

            if not block_rows:
                continue
            t0_rows = [r for r in block_rows if r["time_h"] == 0]
            if not t0_rows or t0_rows[0]["mean_recomputed"] in (None, 0):
                t0_mean = None
                t0_sd = None
            else:
                t0_mean = t0_rows[0]["mean_recomputed"] if recompute_mean else t0_rows[0]["mean_sheet"]
                t0_sd = t0_rows[0]["sd_recomputed"] if recompute_mean else t0_rows[0]["sd_sheet"]

            for r in block_rows:
                mean_value = r["mean_recomputed"] if recompute_mean else r["mean_sheet"]
                sd_value = r["sd_recomputed"] if recompute_mean else r["sd_sheet"]
                r["mean_used"] = mean_value
                r["sd_used"] = sd_value
                r["mean_t0_normalized"] = (mean_value / t0_mean) if (mean_value is not None and t0_mean not in (None, 0)) else None
                r["sd_t0_normalized"] = (sd_value / t0_mean) if (sd_value is not None and t0_mean not in (None, 0)) else None
                records.append(r)

    df = pd.DataFrame.from_records(records)
    if not df.empty:
        df = df.sort_values(["cell_line", "exposure_seconds", "time_h"]).reset_index(drop=True)
    return df


def select_target_vector(
    df: pd.DataFrame,
    *,
    cell_line: str,
    exposure_seconds: int,
    mode: str = "t0_normalized",
    time_points: Iterable[int] = (0, 24, 48, 72),
) -> tuple[pd.DataFrame, pd.Series, pd.Series]:
    """Return rows, target y, and sigma for one cell-line/exposure."""
    rows = df[
        (df["cell_line"].astype(str).str.lower() == cell_line.lower())
        & (df["exposure_seconds"].astype(int) == int(exposure_seconds))
        & (df["time_h"].astype(int).isin([int(t) for t in time_points]))
    ].sort_values("time_h")

    if rows.empty:
        raise ValueError(f"No target rows found for cell_line={cell_line!r}, exposure_seconds={exposure_seconds}")

    if mode == "raw":
        y = rows["mean_used"].astype(float)
        sigma = rows["sd_used"].astype(float)
    elif mode == "t0_normalized":
        y = rows["mean_t0_normalized"].astype(float)
        sigma = rows["sd_t0_normalized"].astype(float)
    else:
        raise ValueError("mode must be 'raw' or 't0_normalized'")

    sigma = sigma.replace(0, pd.NA).fillna(max(0.05, float(y.mean()) * 0.05)).astype(float)
    return rows, y, sigma
